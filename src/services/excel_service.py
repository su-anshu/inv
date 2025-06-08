"""
Excel Service - Handles all Excel file operations and data management
"""

import pandas as pd
import openpyxl
from pathlib import Path
import config
from datetime import datetime, date
from typing import Dict, List, Optional, Any
import logging

class ExcelService:
    """Service class for Excel file operations"""
    
    def __init__(self):
        self.file_path = config.EXCEL_FILE_PATH
        self.sheet_names = config.SHEET_NAMES
        self.logger = logging.getLogger(__name__)
        
    def file_exists(self) -> bool:
        """Check if Excel file exists"""
        try:
            return self.file_path.exists() and self.file_path.is_file()
        except Exception as e:
            self.logger.error(f"Error checking file existence: {str(e)}")
            return False
    
    def get_sheet_names(self) -> List[str]:
        """Get all sheet names from Excel file"""
        try:
            if not self.file_exists():
                return []
            
            with pd.ExcelFile(self.file_path) as xl:
                return xl.sheet_names
        except Exception as e:
            self.logger.error(f"Error getting sheet names: {str(e)}")
            return []
    
    def read_sheet(self, sheet_name: str) -> Optional[pd.DataFrame]:
        """Read data from specific sheet"""
        try:
            if not self.file_exists():
                self.logger.error("Excel file not found")
                return None
            
            df = pd.read_excel(self.file_path, sheet_name=sheet_name)
            self.logger.info(f"Successfully read sheet '{sheet_name}' with {len(df)} rows")
            return df
            
        except Exception as e:
            self.logger.error(f"Error reading sheet '{sheet_name}': {str(e)}")
            return None
    
    def read_stock_data(self) -> Optional[pd.DataFrame]:
        """Read main stock data from the stock sheet"""
        try:
            stock_sheet_name = self.sheet_names.get('STOCK', 'stock sheet')
            df = self.read_sheet(stock_sheet_name)
            
            if df is None:
                return None
            
            # Process and standardize column names
            df = self.standardize_stock_columns(df)
            
            return df
            
        except Exception as e:
            self.logger.error(f"Error reading stock data: {str(e)}")
            return None
    
    def standardize_stock_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """Standardize column names for stock data"""
        try:
            # Create a copy to avoid modifying original
            df_clean = df.copy()
            
            # Common column name mappings (adapt based on your Excel structure)
            column_mappings = {
                'product': 'product_name',
                'weight': 'weight_kg',
                'current stock': 'current_stock',
                'opening stock': 'opening_stock',
                'closing stock': 'closing_stock',
                'minimum stock': 'min_stock',
                'price': 'unit_price',
                'value': 'stock_value'
            }
            
            # Normalize column names (lowercase, remove extra spaces)
            df_clean.columns = df_clean.columns.str.lower().str.strip()
            
            # Apply mappings
            for old_name, new_name in column_mappings.items():
                if old_name in df_clean.columns:
                    df_clean.rename(columns={old_name: new_name}, inplace=True)
            
            # Fill NaN values
            numeric_columns = df_clean.select_dtypes(include=['number']).columns
            df_clean[numeric_columns] = df_clean[numeric_columns].fillna(0)
            
            text_columns = df_clean.select_dtypes(include=['object']).columns
            df_clean[text_columns] = df_clean[text_columns].fillna('')
            
            return df_clean
            
        except Exception as e:
            self.logger.error(f"Error standardizing columns: {str(e)}")
            return df
    
    def write_sheet(self, df: pd.DataFrame, sheet_name: str, mode: str = 'replace') -> bool:
        """Write DataFrame to Excel sheet"""
        try:
            if not self.file_exists():
                # Create new file if it doesn't exist
                with pd.ExcelWriter(self.file_path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                self.logger.info(f"Created new Excel file with sheet '{sheet_name}'")
                return True
            
            if mode == 'replace':
                # Replace entire sheet
                with pd.ExcelWriter(self.file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            elif mode == 'append':
                # Append to existing sheet (this is complex with openpyxl)
                self._append_to_sheet(df, sheet_name)
            
            self.logger.info(f"Successfully wrote to sheet '{sheet_name}'")
            return True
            
        except Exception as e:
            self.logger.error(f"Error writing to sheet '{sheet_name}': {str(e)}")
            return False
    
    def _append_to_sheet(self, df: pd.DataFrame, sheet_name: str):
        """Append data to existing sheet"""
        try:
            # Load existing workbook
            workbook = openpyxl.load_workbook(self.file_path)
            
            if sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Find the next empty row
                next_row = worksheet.max_row + 1
                
                # Write data
                for r_idx, row in enumerate(df.values, start=next_row):
                    for c_idx, value in enumerate(row, start=1):
                        worksheet.cell(row=r_idx, column=c_idx, value=value)
            else:
                # Create new sheet if it doesn't exist
                worksheet = workbook.create_sheet(sheet_name)
                
                # Write headers
                for c_idx, header in enumerate(df.columns, start=1):
                    worksheet.cell(row=1, column=c_idx, value=header)
                
                # Write data
                for r_idx, row in enumerate(df.values, start=2):
                    for c_idx, value in enumerate(row, start=1):
                        worksheet.cell(row=r_idx, column=c_idx, value=value)
            
            # Save workbook
            workbook.save(self.file_path)
            workbook.close()
            
        except Exception as e:
            self.logger.error(f"Error appending to sheet: {str(e)}")
            raise
    
    def record_sale(self, sale_data: Dict[str, Any]) -> bool:
        """Record a new sale transaction"""
        try:
            # Prepare sale record
            sale_record = {
                'Date': sale_data.get('date', datetime.now().date()),
                'Product': sale_data.get('product', ''),
                'Quantity': sale_data.get('quantity', 0),
                'Unit_Price': sale_data.get('price', 0),
                'Total_Amount': sale_data.get('total_amount', 0),
                'Channel': sale_data.get('channel', ''),
                'Order_ID': sale_data.get('order_id', ''),
                'Customer_Name': sale_data.get('customer_name', ''),
                'Timestamp': datetime.now()
            }
            
            # Convert to DataFrame
            sale_df = pd.DataFrame([sale_record])
            
            # Append to sales sheet (you might want to create a separate sales sheet)
            success = self.write_sheet(sale_df, 'Sales_Log', mode='append')
            
            if success:
                # Update stock levels
                self._update_stock_after_sale(sale_data)
            
            return success
            
        except Exception as e:
            self.logger.error(f"Error recording sale: {str(e)}")
            return False
    
    def record_purchase(self, purchase_data: Dict[str, Any]) -> bool:
        """Record a new purchase transaction"""
        try:
            # Prepare purchase record
            purchase_record = {
                'Date': purchase_data.get('date', datetime.now().date()),
                'Supplier': purchase_data.get('supplier', ''),
                'Material_Type': purchase_data.get('material_type', ''),
                'Quantity_KG': purchase_data.get('quantity', 0),
                'Rate_Per_KG': purchase_data.get('rate_per_kg', 0),
                'Total_Amount': purchase_data.get('total_amount', 0),
                'Invoice_Number': purchase_data.get('invoice_number', ''),
                'Quality_Grade': purchase_data.get('quality_grade', ''),
                'Payment_Method': purchase_data.get('payment_method', ''),
                'Notes': purchase_data.get('notes', ''),
                'Timestamp': datetime.now()
            }
            
            # Convert to DataFrame
            purchase_df = pd.DataFrame([purchase_record])
            
            # Append to purchases sheet
            success = self.write_sheet(purchase_df, 'Purchase_Log', mode='append')
            
            return success
            
        except Exception as e:
            self.logger.error(f"Error recording purchase: {str(e)}")
            return False
    
    def record_production(self, production_data: Dict[str, Any]) -> bool:
        """Record production data"""
        try:
            # Prepare production record
            production_record = {
                'Date': production_data.get('date', datetime.now().date()),
                'Batch_Number': production_data.get('batch_number', ''),
                'Raw_Material_Used_KG': production_data.get('raw_material_used', 0),
                'Operator': production_data.get('operator', ''),
                'Shift': production_data.get('shift', ''),
                'Efficiency_Percent': production_data.get('efficiency', 0),
                'Quality_Notes': production_data.get('quality_notes', ''),
                'Issues': production_data.get('issues', ''),
                'Remarks': production_data.get('remarks', ''),
                'Timestamp': datetime.now()
            }
            
            # Add output data for each product
            output_data = production_data.get('output_data', {})
            for product, quantity in output_data.items():
                if quantity > 0:
                    production_record[f'Output_{product}'] = quantity
            
            # Convert to DataFrame
            production_df = pd.DataFrame([production_record])
            
            # Append to production sheet
            success = self.write_sheet(production_df, 'Production_Log', mode='append')
            
            if success:
                # Update stock levels with production output
                self._update_stock_after_production(output_data)
            
            return success
            
        except Exception as e:
            self.logger.error(f"Error recording production: {str(e)}")
            return False
    
    def _update_stock_after_sale(self, sale_data: Dict[str, Any]):
        """Update stock levels after a sale"""
        try:
            # This is a simplified version - you'd implement based on your Excel structure
            product = sale_data.get('product', '')
            quantity_sold = sale_data.get('quantity', 0)
            
            # Read current stock
            stock_df = self.read_stock_data()
            if stock_df is not None:
                # Find the product row and update stock
                # Implementation depends on your Excel structure
                pass
            
        except Exception as e:
            self.logger.error(f"Error updating stock after sale: {str(e)}")
    
    def _update_stock_after_production(self, output_data: Dict[str, int]):
        """Update stock levels after production"""
        try:
            # This is a simplified version - you'd implement based on your Excel structure
            for product, quantity in output_data.items():
                if quantity > 0:
                    # Update stock for this product
                    pass
            
        except Exception as e:
            self.logger.error(f"Error updating stock after production: {str(e)}")
    
    def backup_file(self, backup_path: Path) -> bool:
        """Create backup of Excel file"""
        try:
            if not self.file_exists():
                return False
            
            # Copy file to backup location
            import shutil
            shutil.copy2(self.file_path, backup_path)
            
            self.logger.info(f"Backup created at: {backup_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error creating backup: {str(e)}")
            return False
    
    def get_file_info(self) -> Dict[str, Any]:
        """Get information about the Excel file"""
        try:
            if not self.file_exists():
                return {'exists': False}
            
            stat = self.file_path.stat()
            sheet_names = self.get_sheet_names()
            
            return {
                'exists': True,
                'path': str(self.file_path),
                'size_mb': stat.st_size / (1024 * 1024),
                'modified': datetime.fromtimestamp(stat.st_mtime),
                'sheet_names': sheet_names,
                'sheet_count': len(sheet_names)
            }
            
        except Exception as e:
            self.logger.error(f"Error getting file info: {str(e)}")
            return {'exists': False, 'error': str(e)}